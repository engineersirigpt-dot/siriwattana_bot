"""Turn the Markdown tables in a chat answer into a real .xlsx workbook.

Used by POST /chat/export-xlsx. Each Markdown table in the answer becomes its
own worksheet (header row styled, columns auto-sized, header frozen). If the
answer has no table we still return a one-cell sheet so the download succeeds
with a clear message.
"""

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_SEP_CELL = re.compile(r"^:?-{1,}:?$")


def _split_row(line: str) -> list[str]:
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [c.strip() for c in s.split("|")]


def _is_separator(line: str) -> bool:
    if "|" not in line:
        return False
    cells = [c for c in _split_row(line) if c != ""]
    return len(cells) > 0 and all(_SEP_CELL.match(c) for c in cells)


def parse_markdown_tables(md: str) -> list[list[list[str]]]:
    """Return a list of tables; each table is a list of rows (incl. header)."""
    lines = md.split("\n")
    tables: list[list[list[str]]] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if (
            "|" in line
            and line.strip()
            and i + 1 < len(lines)
            and _is_separator(lines[i + 1])
        ):
            header = _split_row(line)
            rows = [header]
            j = i + 2
            while j < len(lines) and "|" in lines[j] and lines[j].strip():
                cells = _split_row(lines[j])
                # Pad/trim to header width so the grid stays rectangular.
                if len(cells) < len(header):
                    cells += [""] * (len(header) - len(cells))
                rows.append(cells[: len(header)])
                j += 1
            tables.append(rows)
            i = j
        else:
            i += 1
    return tables


def export_tables_to_xlsx(content_md: str, title: str | None = None) -> bytes:
    tables = parse_markdown_tables(content_md or "")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(vertical="top", wrap_text=True)

    if not tables:
        ws = wb.create_sheet("ตาราง")
        ws["A1"] = "ไม่พบตารางในคำตอบนี้"
        ws["A1"].font = Font(italic=True, color="888888")
    else:
        for idx, rows in enumerate(tables, 1):
            name = "ตาราง" if len(tables) == 1 else f"ตาราง {idx}"
            ws = wb.create_sheet(name[:31])
            ncols = len(rows[0]) if rows else 0

            for r, row in enumerate(rows, 1):
                for c in range(1, ncols + 1):
                    val = row[c - 1] if c - 1 < len(row) else ""
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = border
                    cell.alignment = wrap_top
                    if r == 1:
                        cell.fill = header_fill
                        cell.font = header_font

            # Auto-ish column widths from the longest cell (capped).
            for c in range(1, ncols + 1):
                longest = max(
                    (len(str(row[c - 1])) for row in rows if c - 1 < len(row)),
                    default=10,
                )
                ws.column_dimensions[get_column_letter(c)].width = min(
                    max(longest + 2, 12), 60
                )

            if len(rows) > 1:
                ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
